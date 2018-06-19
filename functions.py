import openpyxl
import os

def read_predictions():

    # make a list of all files in a directory
    entries = os.listdir("Round_One_Forms")

    # get length of that list
    num_entries = len(entries)

    # create master dictionary of predictions and tuple of keys
    all_predictions = {}
    keys = []

    # iterate through all files to create dictionary
    for i in range(0,num_entries-1)

        wb = openpyxl.load_workbook(entries[i])
        sheet = wb.get_sheet_by_name('Sheet1')
        predictions = {}
        
        # each dict entry is a two element list with (winner,score)
        for p in range(8,57)
            keykey = sheet.cell(row=p,column=2).value
            predictions[keykey] = ((sheet.cell(row=p,column=3).value,sheet.cell(row=p,column=4).value))
        
        # parse filename to create key with filetype removed
        key = entries[i].replace('.xlsx','')
        keys.append(key)

        # predictions goes in the master dictionary all_predictions
        all_predictions[key] = predictions

    return all_predictions

def generate_standings(all_predictions)
    wb = openpyxl.load_workbook('results.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    results = {}

    for p in range(8,57)
        key = sheet.cell(row=p,column=2).value
        results[key] = ((sheet.cell(row=p,column=3).value,sheet.cell(row=p,column=4).value))

    # loop through results excel file, create list of keys for entries that have result
    all_predictions_keys = all_predictions.keys()
    results_keys = results.keys()

    for key in all_predictions_keys
        temp1 = all_predictions[key]  # grabs an entry
        for key in results_keys     # grabs a match key
            actual_winner = results[key][0].lower()
            actual_score = results[key][1]
            predicted_winner = temp1[key][0].lower()
            predicted_score = temp1[key][1]

    
'''
create page that list all of players predictions and results of those predictions

front page:
    latest match included
    standings (name w/ link, correct result, correct score, points)
    list of matches with result, number correct, number correct scores
'''
